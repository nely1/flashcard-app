from tkinter import *
from tkinter.font import BOLD
from PIL import ImageTk, Image
from openpyxl import *
import os
import random
import re
from tkinter import ttk


#establish the window
root = Tk()
root.title('Anki')
root.geometry("975x500")
root.configure(background= "#00FFFF")



#CLASSES
#_______________________________________________________________________________
#Main menu
class Main_menu:

    def __init__(self):
        self.mylabel = Label(root, text="暗記しろ", bg = "#00FFFF", 
            font = ("Helvetica", 72, BOLD))
        self.test_button = Button(root, text = "テスト", 
            font = ("Helvetica", 16, BOLD), padx = 135, pady = 3, bg = "#03A89E"
            ,command = clicked_test)
        self.add_button = Button(root, text = "単語を追加", 
            font = ("Helvetica", 16, BOLD), padx = 110, pady = 3, bg = "#03A89E"
            , command = clicked_add)
        self.exit_button = Button(root, text = "去る", font = ("Helvetica", 16, 
            BOLD), padx = 139, pady = 3, bg = "#03A89E", command = root.quit)
        self.anki = ImageTk.PhotoImage(Image.open("icon/icon.png"))
        self.anki_img = Label(image = self.anki)
        self.anki = ImageTk.PhotoImage(Image.open("icon/icon.png"))
        self.anki_img = Label(image = self.anki)
        

    def show_main_menu(self):
        self.exit_button.place(x = 50, y = 420)
        self.add_button.place(x = 50, y = 270)
        self.mylabel.pack()
        self.test_button.place(x = 50, y = 120)
        self.anki_img.place(x = 450,y = 120)
        
    
    def hide(self):
        self.exit_button.place_forget()
        self.add_button.place_forget()
        self.mylabel.pack_forget()
        self.test_button.place_forget()
        self.anki_img.place_forget()
    

#add menu
class Add_menu:

    def __init__(self):
        self.back_button = Button(root, text = "戻る", 
            font = ("Helvetica", 16, BOLD), padx = 135, pady = 3, bg = "#03A89E", 
            command = self.clicked_back)
        self.add_word = Button(root, text = "送信",
            font = ("Helvetica", 16, BOLD),padx = 135, pady = 3, bg = "#03A89E",
            command = self.clicked_submit)
        self.word_label = Label(root, text = "言葉", font = ("Helvetica", 24),
            bg = "#00FFFF")
        self.def_label = Label(root, text = "意味",  font = ("Helvetica", 24), 
            bg = "#00FFFF")
        self.enter_word = Entry(root, width = 45,font = ("Helvetica", 24))
        self.enter_def = Entry(root, width = 45,font = ("Helvetica", 24))
        self.duplicate = Label(root, text = '言葉はもう記載された',bg = "#00FFFF", 
            font = ("Helvetica", 20, BOLD), fg = '#995566')
        self.error = Label(root, text = 'EXCELが開かれた',bg = "#00FFFF", 
            font = ("Helvetica", 20, BOLD), fg = '#995566')

        

    def show_add_menu(self):
        self.word_label.place(x = 100, y = 60)
        self.def_label.place(x = 100, y = 230)
        self.add_word.place(x = 500, y = 400)
        self.back_button.place(x = 100, y = 400)
        self.enter_word.place(x = 100, y = 120)
        self.enter_def.place(x = 100, y = 290)

    def hide(self):
        self.word_label.place_forget()
        self.def_label.place_forget()
        self.add_word.place_forget()
        self.back_button.place_forget()
        self.enter_word.place_forget()
        self.enter_def.place_forget()
        self.duplicate.pack_forget()
        self.error.pack_forget()
        self.delete_text()
    
    def delete_text(self):
        self.enter_word.delete(0, len(self.enter_word.get()))
        self.enter_def.delete(0, len(self.enter_def.get()))
    
    def clicked_back(self):
        main_menu.show_main_menu()
        self.hide()
    
    def clicked_submit(self):
        present = False
        f = open_file()
        sheet = f.active
        
        if len(self.enter_word.get()) != 0 and len(self.enter_def.get()) != 0:
            for row in range(1, sheet.max_row + 1):
                if self.enter_word.get() == sheet['A' + str(row)].value:
                    present = True
                    self.duplicate.pack()
                    self.duplicate.after(2000, lambda:self.duplicate.pack_forget())
                    break

            if not present:
                sheet.append([self.enter_word.get(), self.enter_def.get()])
                try:
                    f.save("words.xlsx")
                except PermissionError:
                    self.error.pack()
                    self.error.after(2000, lambda:self.error.pack_forget())
                else:
                    self.delete_text()

        f.close()
    

    
#test menu
class Test_menu:
    
    def __init__(self):
        self.slider = Scale(root, from_ = 1, to = 200, orient = 'horizontal'
        ,sliderlength = 10, length = 600, font = ('Helvetica', 18),
        troughcolor= '#03A89E',bg = '#00FFFF', highlightbackground= '#00FFFF',
        activebackground= '#00FFFF')
        self.questions = Label(root, text = "問題数",  font = ("Helvetica", 24), 
            bg = "#00FFFF")
        self.back = self.back_button = Button(root, text = "戻る", 
            font = ("Helvetica", 16, BOLD), padx = 135, pady = 3, bg = "#03A89E", 
            command = self.clicked_back)
        self.r = IntVar()
        self.eng_jap = Radiobutton(root, text = "english -> 日本語",
        variable = self.r, value = 2, font = ("Helvetica", 24), bg = "#00FFFF")
        self.jap_eng = Radiobutton(root, text = "日本語 -> english", 
        variable = self.r, value = 1, font = ("Helvetica", 24), bg = "#00FFFF")     
        self.start = Button(root, text = "開始", 
        font = ("Helvetica", 16, BOLD), padx = 135, pady = 3, bg = "#03A89E", 
        command = lambda: self.click_start(self.r.get(),self.slider.get()))
        
        
    
    def show_menu(self):
        self.slider.place(x = 220, y = 130)
        self.back.place(x = 100, y = 400)
        self.questions.place(x = 100, y = 150)
        self.start.place(x = 500, y = 400)
        self.eng_jap.place(x = 500, y = 275)
        self.jap_eng.place(x = 100, y = 275)
        self.r.set(1)
        
    
    def hide(self):
        self.slider.place_forget()
        self.questions.place_forget()
        self.start.place_forget()
        self.back.place_forget()
        self.eng_jap.place_forget()
        self.jap_eng.place_forget()
    
    def clicked_back(self):
        self.hide()
        main_menu.show_main_menu()

    def click_start(self, mode, question_num):
        self.hide()
        f = open_file()
        sheet = f.active
        test = Actual_test(question_num, sheet, mode)
        test.do_test()
        f.close()

        
            
    
#actual test

class Actual_test:

    def __init__(self, questions, sheet, mode):
        self.question_text = StringVar()
        self.word_q = Label(root, font = ("Helvetica", 33), 
        textvariable= self.question_text, bg = "#00FFFF")
        self.answer = Entry(root, width = 45, font = ('Helvetica', 24)) 
        self.questions = questions
        self.sheet = sheet
        self.mode = mode
        self.curr_q = 1
        self.next = Button(root, text = "次", 
        font = ("Helvetica", 16, BOLD), padx = 135, pady = 3, bg = "#03A89E", 
        command = self.next_question)
        self.prev = Button(root, text = "以前", 
        font = ("Helvetica", 16, BOLD), padx = 135, pady = 3, bg = "#03A89E", 
        command = self.prev_question)
        self.memory = []
        self.num = StringVar()
        self.q_num = Label(root, font = ("Helvetica", 24), 
        textvariable = self.num, bg = "#00FFFF")
        self.finish = Button(root, text = "完了", 
        font = ("Helvetica", 16, BOLD), padx = 135, pady = 3, bg = "#c19b6c", 
        command = self.finish_click)
        self.back = Button(root, text = "<-戻る", 
            font = ("Helvetica", 16, BOLD), padx = 25, pady = 3, bg = "#9b55ee", 
            command = self.clicked_back)
        self.solutions = []
    
    def next_question(self):
        self.memory[self.curr_q - 1][1] = self.answer.get()
       
        #the case when we are at a new question
        if self.curr_q == len(self.memory):
            self.generate_question()
        #the case when we are at a previous question
        else: 
            self.question_text.set(self.memory[self.curr_q][0])   
        
        self.curr_q += 1
        self.num.set(str(self.curr_q) + " / " + str(self.questions))
        self.answer.delete(0, len(self.answer.get()))
        self.answer.insert(0, self.memory[self.curr_q - 1][1])

        if self.curr_q == self.questions:
            self.next.place_forget()
            self.finish.place(x = 500, y = 400)
            
        
     
    def finish_click(self):
        self.memory[self.curr_q - 1][1] = self.answer.get()
        self.hide()
        result = Results(self.memory, self.solutions)
        result.check_answers()

        

    def clicked_back(self):
        main_menu.show_main_menu()
        self.hide()


    def hide(self):  
        self.finish.place_forget()
        self.answer.place_forget()
        self.prev.place_forget()
        self.q_num.place_forget()
        self.word_q.place_forget()
        self.next.place_forget()
        self.back.place_forget()
        
    
    def prev_question(self):
        if self.curr_q != 1:
            self.curr_q -= 1
            self.question_text.set(self.memory[self.curr_q - 1][0])
            self.num.set(str(self.curr_q) + " / " + str(self.questions))
            self.answer.delete(0, len(self.answer.get()))
            self.answer.insert(0, self.memory[self.curr_q - 1][1])

            if not self.next.winfo_ismapped() and self.finish.winfo_ismapped():
                self.next.place(x = 500, y = 400)
                self.finish.place_forget()
            
            

    def do_test(self):
        if self.questions == 1:
            self.finish.place(x = 500, y = 400)
        else:
            self.next.place(x = 500, y = 400)
        self.answer.place(x = 100, y = 275)
        self.word_q.place(x = 375, y = 150)
        self.prev.place(x = 100, y = 400)
        self.q_num.place(x = 800, y = 20)
        self.back.place(x = 50, y = 20)
        self.num.set(str(self.curr_q) + " / " + str(self.questions))
        self.generate_question()
    
    def generate_question(self):
        row = random.randint(1, self.sheet.max_row)
        if self.mode == 1:
            self.question_text.set(self.sheet['A' + str(row)].value)
            self.solutions.append(self.sheet['B' + str(row)].value)
        else:
            self.question_text.set(self.sheet['B' + str(row)].value)
            self.solutions.append(self.sheet['A' + str(row)].value)
        self.memory.append([self.question_text.get(), ""])
            
 

class Results:

    def __init__(self, memory, solutions):
        self.solutions = solutions
        self.memory = memory
        self.right_wrong = {'right': [], 'wrong': []}
        self.notebook = ttk.Notebook(root, width = 975, height = 500)
        


    def check_answers(self):
        for x in range(len(self.memory)):
            self.check(self.memory[x], self.solutions[x])
        self.notebook.pack()
        self.create_result_screen('right', '正解')
        self.create_result_screen('wrong', "不正解")


    def check(self, answer, solution):
        matches = 0
        ans_split = re.split("\W",answer[1])
        sol_split = re.split("\W", solution)
        sol_split_copy = sol_split.copy()
        for word in ans_split:
            if word in sol_split_copy:
                matches += 1
                sol_split_copy.remove(word)
        if matches / len(sol_split) >= 0.5:
            self.right_wrong['right'].append(answer + [solution])
        else:
            self.right_wrong['wrong'].append(answer + [solution])
    
    def create_result_screen(self, key, heading):
        main_frame = Frame(self.notebook,width = 975, height = 500, bg = '#03A89E')
        main_frame.pack(fill = BOTH, expand = 1)
        self.notebook.add(main_frame, text = heading)

        #title
        Label(main_frame, text = heading, font = ('Helvetica', 24, BOLD), 
        bg = '#03A89E').pack()
        
        #buttons 
        redo = Button(main_frame, text = 'やり直し', padx = 25, pady = 2,
            bg = '#cbc07e', font = ("Helvetica", 14), command = self.click_redo)
        redo.place(x = 40,  y = 5)
        go_back = Button(main_frame, text = "戻る", padx = 25, pady = 2, 
            bg= '#cbc07e', font = ("Helvetica", 14), 
                command = self.click_go_back)
        go_back.place(x = 825, y = 5)
        
        #display the results in a table
        self.show_results(key, heading, main_frame)
    
    def click_go_back(self):
        self.notebook.pack_forget()
        main_menu.show_main_menu()
    
    def click_redo(self):
        self.notebook.pack_forget()
        test.show_menu()
        

    def show_results(self, answers, heading, place):
        style = ttk.Style()
        style.theme_use("clam")
        #configure treeview colours
        style.configure("Treeview", 
            background = "#e6f7fd",
            rowheight = 40,
            fieldbackground = "#ffffff",
            font = 10    
        )
        style.map('Treeview', 
            background = [('selected', '#7eb0cb')])
        
        #create the scrollbars
        tree_frame = Frame(place)
        tree_frame.pack(pady = 20)
        vert_scroll = ttk.Scrollbar(tree_frame)
        vert_scroll.pack(side = RIGHT,fill = Y)
        ho_scroll = ttk.Scrollbar(tree_frame, orient = 'horizontal')
        ho_scroll.pack(side = BOTTOM, fill = X)

        
        #initialize the treeview
        tree = ttk.Treeview(tree_frame, 
            yscrollcommand = vert_scroll.set,
            xscrollcommand = ho_scroll.set,
            selectmode = 'browse'        
        )
        tree.pack(anchor = W, padx = 40, pady = 20, side = BOTTOM)
      
        #configure the scrollbars
        ho_scroll.config(command = tree.xview)
        vert_scroll.config(command = tree.yview)
      
        #define columns for treeview
        tree['columns'] = ('問題', '答えた', '回答')

        #format columns
        tree.column('#0', width = 0, stretch = NO)
        tree.column('問題', anchor = W, width = 300)
        tree.column('答えた', anchor = CENTER, width = 300)
        tree.column('回答', anchor = W, width = 300)

        #create headings
        tree.heading("#0", text = "", anchor = W)
        tree.heading('問題', text = '問題', anchor = W)
        tree.heading('答えた', text = '答えた', anchor = CENTER)
        tree.heading('回答', text = '回答', anchor = W)

        #striped rows
        if answers == 'right':
            tree.tag_configure('oddrow', background = '#9cff8c') 
        else:
            tree.tag_configure('oddrow', background = '#f47b7b')
        tree.tag_configure('evenrow', background = '#ffffff')

        #add data
        count = 1
        for i in self.right_wrong[answers]:
            if count % 2 == 0:
                tree.insert(parent = '', index = 'end', iid = count, text = "",
                    values =(i[0], i[1], i[2]), tags = ('evenrow',))
            else:
                tree.insert(parent = '', index = 'end', iid = count, text = "",
                    values =(i[0], i[1], i[2]), tags = ('oddrow',))
            count += 1
        

#FUNCTIONS
#_______________________________________________________________________________
def main():
    main_menu.show_main_menu()
  
def clicked_test():
    test.show_menu()
    main_menu.hide()



def clicked_add():
    main_menu.hide()
    add.show_add_menu()
    


def open_file():
    name = "words.xlsx"
    if not os.path.exists(name):
        workbook = Workbook()
        workbook.save(filename = name)

    workbook = load_workbook(name)
    return workbook


#global variables
main_menu = Main_menu()
add = Add_menu()
test = Test_menu()



if __name__ == "__main__":
    main()


root.mainloop()